#!/usr/bin/env python3
# coding=utf-8

import config as cfg
import openpyxl
from report_base import ReportBase


OVERVIEW_SHEET_NAME = 'Overview'
STAGE_STATUS_YES = "√"
STAGE_STATUS_PASS = "pass"
STAGE_STATUS_FAIL = "fail"
THICK = openpyxl.styles.Side(style="thin", color='AAAAAA')
# Style
STYLE_DEFAULT = openpyxl.styles.NamedStyle(
    'Default',
    font=openpyxl.styles.Font(sz=12),
    fill=openpyxl.styles.PatternFill('solid', 'EEEEEE'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('left', 'center')
)
STYLE_TITLE = openpyxl.styles.NamedStyle(
    'Title',
    font=openpyxl.styles.Font(sz=12, b=True),
    fill=openpyxl.styles.PatternFill('solid', 'FAF395'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('center', 'center')
)
STYLE_MENU_ARR = [
    openpyxl.styles.NamedStyle(
        'Menu_1',
        font=openpyxl.styles.Font(sz=12, b=True),
        fill=openpyxl.styles.PatternFill('solid', 'B6E0FA'),
        border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
        alignment=openpyxl.styles.Alignment('center', 'center')
    ),
    openpyxl.styles.NamedStyle(
        'Menu_2',
        font=openpyxl.styles.Font(sz=12, b=True),
        fill=openpyxl.styles.PatternFill('solid', 'B6FAC6'),
        border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
        alignment=openpyxl.styles.Alignment('center', 'center')
    ),
]



STYLE_PASS = openpyxl.styles.NamedStyle(
    'Pass',
    font=openpyxl.styles.Font(sz=12, b=True),
    fill=openpyxl.styles.PatternFill('solid', '66CC00'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('center', 'center')
)
STYLE_FAIL = openpyxl.styles.NamedStyle(
    'Fail',
    font=openpyxl.styles.Font(sz=12, b=True),
    fill=openpyxl.styles.PatternFill('solid', 'FF6666'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('center', 'center')
)
STYLE_TOTAL = openpyxl.styles.NamedStyle(
    'Total',
    font=openpyxl.styles.Font(sz=12, b=True),
    fill=openpyxl.styles.PatternFill('solid', '66CCFF'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('center', 'center')
)
STYLE_FINISH = openpyxl.styles.NamedStyle(
    'Finish',
    font=openpyxl.styles.Font(sz=12, b=True),
    fill=openpyxl.styles.PatternFill('solid', '6699FF'),
    border=openpyxl.styles.Border(THICK, THICK, THICK, THICK),
    alignment=openpyxl.styles.Alignment('center', 'center')
)




class ReportExcel(ReportBase):
    def __init__(self, json_out_dir: str, excel_path: str):
        super().__init__(json_out_dir)
        self.excel_path = excel_path
        self.wb = openpyxl.Workbook()
        self.curr_ws = self.wb['Sheet']
        self.curr_row = 2
        self.last_row = 2
        self.last_menu = ''
        self.menu_cnt = 0
        self.max_cols_width = {}
        self.page_idx = 0


    def _before_run(self):
        self.wb = openpyxl.Workbook()
        self.curr_ws = self.wb['Sheet']
        self.curr_row = 2
        self.last_row = 2
        self.last_menu = ''
        self.menu_cnt = 0
        self.max_cols_width = {}
        self.page_idx = 0
        self.wb['Sheet'].title = OVERVIEW_SHEET_NAME
        self._gen_overview_sheet(self.wb[OVERVIEW_SHEET_NAME])


    def _add_menu(self, menu_path: str):
        path_split = menu_path.split('/')
        if len(path_split) < 2:
            return
        if len(path_split) == 2:
            self._gen_case_type_menu()
            self.curr_row = 2
            self.last_row = 2
            self.last_menu = ''
            self.menu_cnt = 0
            self.max_cols_width = {}

            self.curr_ws = self.wb.create_sheet(path_split[1])
            print('Generate sheet {}'.format(path_split[1]))
            self.curr_ws.freeze_panes = 'C2'
            for i, title in enumerate(['case_type', 'case_name'] + #, 'script path'] +
                                    [stage for _, stage in cfg.STAGE_DEFINE.items()]):
                col = 1 + i
                c = self.curr_ws.cell(1, col)
                c.value = title
                c.style = STYLE_TITLE
                self._cal_max_cols_width(title, col)
            self._add_overview_line(self.wb[OVERVIEW_SHEET_NAME], path_split[1], self.page_idx, 3)
            self.page_idx += 1
        self._gen_case_type_menu()
        self.last_menu = menu_path
        self.last_row  = self.curr_row


    def _add_case(self, menu_path: str, file_name: str, case_name: str, result_dict: dict):
        self.curr_ws.cell(self.curr_row, 2).value = case_name
        self.curr_ws.cell(self.curr_row, 2).style = STYLE_DEFAULT
        self._cal_max_cols_width(case_name, 2)

        #self.curr_ws.cell(self.curr_row, 3).value = file_name
        #self.curr_ws.cell(self.curr_row, 3).style = STYLE_DEFAULT
        #self.cal_max_cols_width(file_name, 3)
        for i, (stage_key, stage) in enumerate(cfg.STAGE_DEFINE.items()):
            c = self.curr_ws.cell(self.curr_row, 3 + i)
            if result_dict[stage_key] == self.CASE_RESULT_PASS:
                c.value = STAGE_STATUS_PASS
                c.style = STYLE_PASS
            elif result_dict[stage_key] == self.CASE_RESULT_FAIL:
                c.value = STAGE_STATUS_FAIL
                c.style = STYLE_FAIL
            elif result_dict[stage_key] == self.CASE_RESULT_NO_TEST:
                c.value = STAGE_STATUS_YES
                c.style = STYLE_TOTAL
            else:
                c.style = STYLE_DEFAULT
            self._cal_max_cols_width(c.value, 3 + i)
        self.curr_row += 1


    def _after_run(self):
        self._gen_case_type_menu()
        self.curr_ws.freeze_panes = '{}2'.format(chr(ord('A') + 2))
        self._add_overview_sum_line(self.wb[OVERVIEW_SHEET_NAME], self.page_idx)
        self.wb.save(self.excel_path)


    def _gen_overview_sheet(self, ws: openpyxl.worksheet):
        print('Generate sheet > Overview...')
        ws.merge_cells('A1:A2')
        ws.freeze_panes = 'B3'
        ws.column_dimensions['A'].width = 30
        c = ws.cell(1, 1)
        c.value = '类型'
        c.style = STYLE_TITLE
        for i, (_, stage) in enumerate(cfg.STAGE_DEFINE.items()):
            c = ws.cell(1, i * 4 + 2)
            ws.merge_cells(None, 1, i * 4 + 2, 1, i * 4 + 5)
            c.value = stage
            c.style = STYLE_TITLE

            for j, (title, style) in enumerate(
                    zip(('Pass', 'Fail', 'total', '完成率'),
                        (STYLE_PASS, STYLE_FAIL, STYLE_TOTAL, STYLE_FINISH))):
                c = ws.cell(2, 2 + i * 4 + j)
                c.value = title
                c.style = style


    def _add_overview_line(self, ws: openpyxl.worksheet, page_name: str, page_idx: int,
                        stage_col: int):
        ws.cell(page_idx + 3, 1).value = page_name
        ws.cell(page_idx + 3, 1).hyperlink = ('#{}!A1'.format(page_name))
        ws.cell(page_idx + 3, 1).style = STYLE_DEFAULT
        for i, (_, _) in enumerate(cfg.STAGE_DEFINE.items()):
            col_str = chr(ord('A') + stage_col + i - 1)
            countif_pass = 'COUNTIF(INDIRECT("{}!{}:{}"), "{}")'\
                .format(page_name, col_str, col_str, STAGE_STATUS_PASS)
            countif_fail = 'COUNTIF(INDIRECT("{}!{}:{}"), "{}")'\
                .format(page_name, col_str, col_str, STAGE_STATUS_FAIL)
            countif_yes = 'COUNTIF(INDIRECT("{}!{}:{}"), "{}")'\
                .format(page_name, col_str, col_str, STAGE_STATUS_YES)

            pass_cell = ws.cell(page_idx + 3, 2 + i * 4 + 0)
            fail_cell = ws.cell(page_idx + 3, 2 + i * 4 + 1)
            total_cell = ws.cell(page_idx + 3, 2 + i * 4 + 2)
            finish_cell = ws.cell(page_idx + 3, 2 + i * 4 + 3)

            pass_cell.value = '=' + countif_pass
            fail_cell.value = '=' + countif_fail
            total_cell.value = '=' + countif_yes\
                + '+{}+{}'.format(pass_cell.coordinate, fail_cell.coordinate)
            finish_cell.value = '={}/IF({}=0, 1, {})'\
                .format(pass_cell.coordinate, total_cell.coordinate,
                        total_cell.coordinate)

            pass_cell.style = STYLE_DEFAULT
            fail_cell.style = STYLE_DEFAULT
            total_cell.style = STYLE_DEFAULT
            finish_cell.style = STYLE_DEFAULT

            finish_cell.number_format = '0.00%'


    def _add_overview_sum_line(self, ws: openpyxl.worksheet, page_num: int):
        sum_row = 3 + page_num
        start_row = 3
        end_row = 3 + page_num - 1
        ws.cell(sum_row, 1).value = 'All'
        ws.cell(sum_row, 1).style = STYLE_TITLE
        for i, (_, _) in enumerate(cfg.STAGE_DEFINE.items()):
            pass_cell = ws.cell(sum_row, 2 + i * 4 + 0)
            fail_cell = ws.cell(sum_row, 2 + i * 4 + 1)
            total_cell = ws.cell(sum_row, 2 + i * 4 + 2)
            finish_cell = ws.cell(sum_row, 2 + i * 4 + 3)

            pass_cell.value = '=SUM({}:{})'\
                .format(ws.cell(start_row, pass_cell.col_idx).coordinate,
                        ws.cell(end_row, pass_cell.col_idx).coordinate)
            fail_cell.value = '=SUM({}:{})'\
                .format(ws.cell(start_row, fail_cell.col_idx).coordinate,
                        ws.cell(end_row, fail_cell.col_idx).coordinate)
            total_cell.value = '=SUM({}:{})'\
                .format(ws.cell(start_row, total_cell.col_idx).coordinate,
                        ws.cell(end_row, total_cell.col_idx).coordinate)
            finish_cell.value = '={}/IF({}=0, 1, {})'\
                .format(pass_cell.coordinate, total_cell.coordinate,
                        total_cell.coordinate)

            pass_cell.style = STYLE_DEFAULT
            fail_cell.style = STYLE_DEFAULT
            total_cell.style = STYLE_DEFAULT
            finish_cell.style = STYLE_DEFAULT

            finish_cell.number_format = '0.00%'


    def _gen_case_type_menu(self):
        if self.last_menu != '' and self.curr_row != self.last_row:
            c = self.curr_ws.cell(self.last_row, 1)
            c.value = self.last_menu
            c.style = STYLE_MENU_ARR[self.menu_cnt % len(STYLE_MENU_ARR)]
            self.menu_cnt  += 1
            self._cal_max_cols_width(self.last_menu, 1)
            self.curr_ws.merge_cells(None, self.last_row, 1, self.curr_row - 1, 1)


    def _cal_max_cols_width(self, value: str, col: int):
        if self.max_cols_width.get(col) is None:
            self.max_cols_width[col] = len(value)
            self.curr_ws.column_dimensions[chr(ord('A') + col - 1)].width = self.max_cols_width[col]
        else:
            self.max_cols_width[col] = max(self.max_cols_width[col], len(value))
            self.curr_ws.column_dimensions[chr(ord('A') + col - 1)].width = self.max_cols_width[col]


if __name__ == '__main__':
    report = ReportExcel('./pipeline_i6f/', 'output.xlsx')
    report.run()
    pass

